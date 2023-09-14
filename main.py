import requests
import time
import openpyxl
from multiprocessing import Process

COOKIES_OFFER = {
    'visitor-id': '51df3a40-42f1-4268-ad6a-8f581650054a',
    '_ym_uid': '1693941113102369760',
    '_ym_d': '1693941113',
    '_ga': 'GA1.1.1739919803.1693941113',
    'siteversion': '2',
    # 'yandex-geolocation': '%7B%22name%22%3A%22%D0%A1%D0%B5%D1%80%D0%B3%D0%B8%D0%B5%D0%B2%20%D0%9F%D0%BE%D1%81%D0%B0%D0%B4%22%2C%22description%22%3A%22%D0%9C%D0%BE%D1%81%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%B0%D1%8F%20%D0%BE%D0%B1%D0%BB%D0%B0%D1%81%D1%82%D1%8C%22%2C%22position%22%3A%7B%22latitude%22%3A56.3044%2C%22longitude%22%3A38.1265%7D%7D',
    'last-location': '37149',
    'smartbe-banner': 'smartbe_a',
    'map-center': '%5B55.8188%2C37.7108%5D',
    'map-zoom': '10',
    'current-region-long-lat': '%257B%2522id%2522%253A2504%252C%2522name%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522emexName%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522longitude%2522%253A37.620393%252C%2522latitude%2522%253A55.75396%252C%2522areaId%2522%253A2504%252C%2522areaName%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522countryId%2522%253A3%252C%2522countryName%2522%253A%2522%25D0%25A0%25D0%25BE%25D1%2581%25D1%2581%25D0%25B8%25D1%258F%2522%252C%2522bestLocationId%2522%253A22096%257D',
    '_hjSessionUser_2159954': 'eyJpZCI6ImJmNTZmMWMzLTBmZmMtNWJmMS1iNzJkLWE5NDkzNjY3ODUzMCIsImNyZWF0ZWQiOjE2OTM5NDExMTM1MTUsImV4aXN0aW5nIjp0cnVlfQ==',
    'show-smartbe-widget': 'false',
    'NSC_xxx.fnfy.sv': 'ffffffffc3a01c1e45525d5f4f58455e445a4a42d8a4',
    '_hjIncludedInSessionSample_2159954': '0',
    '_hjSession_2159954': 'eyJpZCI6ImYwZTUzNzg5LTJiMTAtNGI4Zi1iOWMzLWMyYWZhMGE1MTY2NSIsImNyZWF0ZWQiOjE2OTQxMTM3MzQ2NjMsImluU2FtcGxlIjpmYWxzZX0=',
    '_hjAbsoluteSessionInProgress': '0',
    '_hjHasCachedUserAttributes': 'true',
    '_ym_isad': '2',
    '_ga_ELJBTQE1JD': 'GS1.1.1694113735.4.1.1694113810.0.0.0',
    'ph_phc_TVmxW5q4obIgxwCeFEAfJzVxTS8B1NSCEv9Q19FmUdm_posthog': '%7B%22distinct_id%22%3A%2218a66c21e77fa5-0a0df937d11d5a-145a3874-144000-18a66c21e78135a%22%2C%22%24device_id%22%3A%2218a66c21e77fa5-0a0df937d11d5a-145a3874-144000-18a66c21e78135a%22%2C%22site_version%22%3A%22new%22%2C%22%24initial_referrer%22%3A%22%24direct%22%2C%22%24initial_referring_domain%22%3A%22%24direct%22%2C%22%24referrer%22%3A%22https%3A%2F%2Femex.ru%2Fproducts%2F13028AD20A%2FNissan%2F37149%22%2C%22%24referring_domain%22%3A%22emex.ru%22%2C%22%24sesid%22%3A%5B1694113856770%2C%2218a710c1c4dbd1-03b8bd5330b378-145a3874-144000-18a710c1c4e170c%22%5D%2C%22%24session_recording_enabled%22%3Afalse%2C%22%24active_feature_flags%22%3A%5B%5D%2C%22%24enabled_feature_flags%22%3A%7B%7D%7D',
}

HEADERS_OFFER = {
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'ru,en;q=0.9',
    'Access-Control-Allow-Credentials': 'true',
    'Access-Control-Allow-Origin': 'https://emex.ru',
    'Connection': 'keep-alive',
    # 'Cookie': 'visitor-id=51df3a40-42f1-4268-ad6a-8f581650054a; _ym_uid=1693941113102369760; _ym_d=1693941113; _ga=GA1.1.1739919803.1693941113; siteversion=2; yandex-geolocation=%7B%22name%22%3A%22%D0%A1%D0%B5%D1%80%D0%B3%D0%B8%D0%B5%D0%B2%20%D0%9F%D0%BE%D1%81%D0%B0%D0%B4%22%2C%22description%22%3A%22%D0%9C%D0%BE%D1%81%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%B0%D1%8F%20%D0%BE%D0%B1%D0%BB%D0%B0%D1%81%D1%82%D1%8C%22%2C%22position%22%3A%7B%22latitude%22%3A56.3044%2C%22longitude%22%3A38.1265%7D%7D; last-location=37149; smartbe-banner=smartbe_a; map-center=%5B55.8188%2C37.7108%5D; map-zoom=10; current-region-long-lat=%257B%2522id%2522%253A2504%252C%2522name%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522emexName%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522longitude%2522%253A37.620393%252C%2522latitude%2522%253A55.75396%252C%2522areaId%2522%253A2504%252C%2522areaName%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522countryId%2522%253A3%252C%2522countryName%2522%253A%2522%25D0%25A0%25D0%25BE%25D1%2581%25D1%2581%25D0%25B8%25D1%258F%2522%252C%2522bestLocationId%2522%253A22096%257D; _hjSessionUser_2159954=eyJpZCI6ImJmNTZmMWMzLTBmZmMtNWJmMS1iNzJkLWE5NDkzNjY3ODUzMCIsImNyZWF0ZWQiOjE2OTM5NDExMTM1MTUsImV4aXN0aW5nIjp0cnVlfQ==; show-smartbe-widget=false; NSC_xxx.fnfy.sv=ffffffffc3a01c1e45525d5f4f58455e445a4a42d8a4; _hjIncludedInSessionSample_2159954=0; _hjSession_2159954=eyJpZCI6ImYwZTUzNzg5LTJiMTAtNGI4Zi1iOWMzLWMyYWZhMGE1MTY2NSIsImNyZWF0ZWQiOjE2OTQxMTM3MzQ2NjMsImluU2FtcGxlIjpmYWxzZX0=; _hjAbsoluteSessionInProgress=0; _hjHasCachedUserAttributes=true; _ym_isad=2; _ga_ELJBTQE1JD=GS1.1.1694113735.4.1.1694113810.0.0.0; ph_phc_TVmxW5q4obIgxwCeFEAfJzVxTS8B1NSCEv9Q19FmUdm_posthog=%7B%22distinct_id%22%3A%2218a66c21e77fa5-0a0df937d11d5a-145a3874-144000-18a66c21e78135a%22%2C%22%24device_id%22%3A%2218a66c21e77fa5-0a0df937d11d5a-145a3874-144000-18a66c21e78135a%22%2C%22site_version%22%3A%22new%22%2C%22%24initial_referrer%22%3A%22%24direct%22%2C%22%24initial_referring_domain%22%3A%22%24direct%22%2C%22%24referrer%22%3A%22https%3A%2F%2Femex.ru%2Fproducts%2F13028AD20A%2FNissan%2F37149%22%2C%22%24referring_domain%22%3A%22emex.ru%22%2C%22%24sesid%22%3A%5B1694113856770%2C%2218a710c1c4dbd1-03b8bd5330b378-145a3874-144000-18a710c1c4e170c%22%5D%2C%22%24session_recording_enabled%22%3Afalse%2C%22%24active_feature_flags%22%3A%5B%5D%2C%22%24enabled_feature_flags%22%3A%7B%7D%7D',
    'Referer': 'https://emex.ru/products/1744977/Ford/37149',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 YaBrowser/23.7.4.971 Yowser/2.5 Safari/537.36',
    'X-KL-Ajax-Request': 'Ajax_Request',
    'cache-control': 'no-cache',
    'expires': '0',
    'pragma': 'no-cache',
    'sec-ch-ua': '"Not.A/Brand";v="8", "Chromium";v="114", "YaBrowser";v="23"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'traceparent': '00-3399a2e6eec32c896c0882cec54824c3-33eec7d123a13aac-01',
}

COOKIES_RATING = {
    'visitor-id': '51df3a40-42f1-4268-ad6a-8f581650054a',
    '_ym_uid': '1693941113102369760',
    '_ym_d': '1693941113',
    '_ga': 'GA1.1.1739919803.1693941113',
    'siteversion': '2',
    # 'yandex-geolocation': '%7B%22name%22%3A%22%D0%A1%D0%B5%D1%80%D0%B3%D0%B8%D0%B5%D0%B2%20%D0%9F%D0%BE%D1%81%D0%B0%D0%B4%22%2C%22description%22%3A%22%D0%9C%D0%BE%D1%81%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%B0%D1%8F%20%D0%BE%D0%B1%D0%BB%D0%B0%D1%81%D1%82%D1%8C%22%2C%22position%22%3A%7B%22latitude%22%3A56.3044%2C%22longitude%22%3A38.1265%7D%7D',
    'last-location': '37149',
    'smartbe-banner': 'smartbe_a',
    'map-center': '%5B55.8188%2C37.7108%5D',
    'map-zoom': '10',
    'current-region-long-lat': '%257B%2522id%2522%253A2504%252C%2522name%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522emexName%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522longitude%2522%253A37.620393%252C%2522latitude%2522%253A55.75396%252C%2522areaId%2522%253A2504%252C%2522areaName%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522countryId%2522%253A3%252C%2522countryName%2522%253A%2522%25D0%25A0%25D0%25BE%25D1%2581%25D1%2581%25D0%25B8%25D1%258F%2522%252C%2522bestLocationId%2522%253A22096%257D',
    '_hjSessionUser_2159954': 'eyJpZCI6ImJmNTZmMWMzLTBmZmMtNWJmMS1iNzJkLWE5NDkzNjY3ODUzMCIsImNyZWF0ZWQiOjE2OTM5NDExMTM1MTUsImV4aXN0aW5nIjp0cnVlfQ==',
    'show-smartbe-widget': 'false',
    'NSC_xxx.fnfy.sv': 'ffffffffc3a01c6345525d5f4f58455e445a4a42d8a4',
    '_ym_isad': '2',
    '_ga_ELJBTQE1JD': 'GS1.1.1694629463.7.1.1694629770.0.0.0',
    'ph_phc_TVmxW5q4obIgxwCeFEAfJzVxTS8B1NSCEv9Q19FmUdm_posthog': '%7B%22distinct_id%22%3A%2218a66c21e77fa5-0a0df937d11d5a-145a3874-144000-18a66c21e78135a%22%2C%22%24device_id%22%3A%2218a66c21e77fa5-0a0df937d11d5a-145a3874-144000-18a66c21e78135a%22%2C%22site_version%22%3A%22new%22%2C%22%24initial_referrer%22%3A%22%24direct%22%2C%22%24initial_referring_domain%22%3A%22%24direct%22%2C%22%24referrer%22%3A%22https%3A%2F%2Femex.ru%2Fproducts%2F13028AD20A%2FNissan%2F37149%22%2C%22%24referring_domain%22%3A%22emex.ru%22%2C%22%24sesid%22%3A%5B1694629824534%2C%2218a8fc975ddab9-04810311233b27-135e387a-144000-18a8fc975de4f%22%5D%2C%22%24session_recording_enabled%22%3Afalse%2C%22%24active_feature_flags%22%3A%5B%5D%2C%22%24enabled_feature_flags%22%3A%7B%7D%7D',
}

HEADERS_RATING = {
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'ru,en;q=0.9',
    'Access-Control-Allow-Credentials': 'true',
    'Access-Control-Allow-Origin': 'https://emex.ru',
    'Connection': 'keep-alive',
    # 'Cookie': 'visitor-id=51df3a40-42f1-4268-ad6a-8f581650054a; _ym_uid=1693941113102369760; _ym_d=1693941113; _ga=GA1.1.1739919803.1693941113; siteversion=2; yandex-geolocation=%7B%22name%22%3A%22%D0%A1%D0%B5%D1%80%D0%B3%D0%B8%D0%B5%D0%B2%20%D0%9F%D0%BE%D1%81%D0%B0%D0%B4%22%2C%22description%22%3A%22%D0%9C%D0%BE%D1%81%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%B0%D1%8F%20%D0%BE%D0%B1%D0%BB%D0%B0%D1%81%D1%82%D1%8C%22%2C%22position%22%3A%7B%22latitude%22%3A56.3044%2C%22longitude%22%3A38.1265%7D%7D; last-location=37149; smartbe-banner=smartbe_a; map-center=%5B55.8188%2C37.7108%5D; map-zoom=10; current-region-long-lat=%257B%2522id%2522%253A2504%252C%2522name%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522emexName%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522longitude%2522%253A37.620393%252C%2522latitude%2522%253A55.75396%252C%2522areaId%2522%253A2504%252C%2522areaName%2522%253A%2522%25D0%259C%25D0%25BE%25D1%2581%25D0%25BA%25D0%25B2%25D0%25B0%2522%252C%2522countryId%2522%253A3%252C%2522countryName%2522%253A%2522%25D0%25A0%25D0%25BE%25D1%2581%25D1%2581%25D0%25B8%25D1%258F%2522%252C%2522bestLocationId%2522%253A22096%257D; _hjSessionUser_2159954=eyJpZCI6ImJmNTZmMWMzLTBmZmMtNWJmMS1iNzJkLWE5NDkzNjY3ODUzMCIsImNyZWF0ZWQiOjE2OTM5NDExMTM1MTUsImV4aXN0aW5nIjp0cnVlfQ==; show-smartbe-widget=false; NSC_xxx.fnfy.sv=ffffffffc3a01c6345525d5f4f58455e445a4a42d8a4; _ym_isad=2; _ga_ELJBTQE1JD=GS1.1.1694629463.7.1.1694629770.0.0.0; ph_phc_TVmxW5q4obIgxwCeFEAfJzVxTS8B1NSCEv9Q19FmUdm_posthog=%7B%22distinct_id%22%3A%2218a66c21e77fa5-0a0df937d11d5a-145a3874-144000-18a66c21e78135a%22%2C%22%24device_id%22%3A%2218a66c21e77fa5-0a0df937d11d5a-145a3874-144000-18a66c21e78135a%22%2C%22site_version%22%3A%22new%22%2C%22%24initial_referrer%22%3A%22%24direct%22%2C%22%24initial_referring_domain%22%3A%22%24direct%22%2C%22%24referrer%22%3A%22https%3A%2F%2Femex.ru%2Fproducts%2F13028AD20A%2FNissan%2F37149%22%2C%22%24referring_domain%22%3A%22emex.ru%22%2C%22%24sesid%22%3A%5B1694629824534%2C%2218a8fc975ddab9-04810311233b27-135e387a-144000-18a8fc975de4f%22%5D%2C%22%24session_recording_enabled%22%3Afalse%2C%22%24active_feature_flags%22%3A%5B%5D%2C%22%24enabled_feature_flags%22%3A%7B%7D%7D',
    'Referer': 'https://emex.ru/products/1744977/Ford/37149',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 YaBrowser/23.7.5.704 Yowser/2.5 Safari/537.36',
    'X-KL-Ajax-Request': 'Ajax_Request',
    'cache-control': 'no-cache',
    'expires': '0',
    'pragma': 'no-cache',
    'sec-ch-ua': '"Not.A/Brand";v="8", "Chromium";v="114", "YaBrowser";v="23"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'traceparent': '00-53afc6d5a057cf1670cc1b542d612f6e-034e016206bf4d4c-01',
}


def get_data_request_for_pars(path_to_file: str, number_string: int) -> tuple:
    """Коннект к файлу exel
       path_to_file -> путь до файла
       number_string -> номер строки с нужной парой
       NB! если оставлять exel, то метод лучше вынести для возможности фоновой работы, потому что коннект выполняется
       порядка 8 секунд"""
    wb = openpyxl.load_workbook(path_to_file)
    sheet = wb['Лист1']
    print('connected to db')
    return sheet.cell(number_string, 1).value, sheet.cell(number_string, 2).value


def set_params_offer(make: str, detail_num: str) -> dict:
    """Устанавливает параметры для get-запроса
       make - модель авто
       detail_num - номер детали"""
    params_offer = {
        'make': make,
        'detailNum': detail_num,
        'locationId': '37149',
        'showAll': 'true',
        'longitude': '37.7108',
        'latitude': '55.8188',
    }
    return params_offer


def set_params_rating(offer_key: str) -> dict:
    """Устанавливает параметры для get-запроса по информации о поставщике
       offer_key - идентификатор поставщика"""
    params_rating_offer = {
        'offerKey': offer_key,
    }
    return params_rating_offer


# Генератор большого количества запросов (на каждого поставщика). Кладет всю систему
def get_rating_and_provider_offer(offer_key: str) -> list:
    """Получает рейтинг соответствующего поставщика
       offer_key - идентификатор поставщика"""
    response_rating_offer = requests.get('https://emex.ru/api/search/rating', params=set_params_rating(offer_key),
                                         cookies=COOKIES_RATING, headers=HEADERS_RATING).json()
    return response_rating_offer.get('priceLogo')


def function_for_multiprocessing(list_res_: list, count_, item_):
    """Собирает данные и закидывает в список. Также инициирует get-запросы по поставщикам.
       Вынесена для создания процессов"""
    list_res_.append(
        f'Номер позиции - {count_}), логотип поставщика: {get_rating_and_provider_offer(item_.get("offerKey"))}, '
        f'рейтинг: {item_.get("rating")}, количество товаров: {item_.get("quantity")}, '
        f'срок доставки: {item_.get("delivery").get("value")}, стоимость товара: '
        f'{item_.get("displayPrice").get("value")}')


# Если в куках оставить геолокацию, то time.sleep достаточно 0,01-0,1с. Этого чаще всего хватает, чтобы запросы прошли
# без блока. Чтобы своё гео не палить, я это закомментировал (в куках), тогда стабильно работает при 0,2с
# но нужно учитывать, что реальная пауза немного больше, порядка 0,3с
#
# NB! если включить мультипроцессинг, то стабильно работает без sleep. 200 запросов + коннект за 31с
def get_data(tuple_data: tuple, search_depth: int):
    """Главный метод. Получает на вход кортеж с ключом и глубину поиска.
       Непосредственно парсер.
       tuple_data - ключ поиска, получает из get_data_request_for_pars
       search_depth - глубина"""
    detail_num, make = tuple_data[0], tuple_data[1]
    response_offer = requests.get('https://emex.ru/api/search/search', params=set_params_offer(make, detail_num),
                                  cookies=COOKIES_OFFER, headers=HEADERS_OFFER).json()
    offers = response_offer.get('searchResult').get('originals')[0].get('offers')
    total_offers = len(offers)

    list_process = []
    list_res = [f'Всего предложений: {total_offers}']
    count = 0
    print(f'Всего предложений: {total_offers}')
    for i in range(search_depth if search_depth < total_offers else total_offers):
        print(count)
        item = offers[i]
        process = Process(target=function_for_multiprocessing, args=(list_res, count, item))
        list_process.append(process)
        process.start()
        count += 1
        # time.sleep(0.2)

    for proc in list_process:
        proc.join()

    with open('result.txt', 'w') as f:
        f.write('_'.join(list_res))
    print('successfully')


def main():
    get_data(get_data_request_for_pars('dict_data_req.xlsx', 3), 200)


if __name__ == '__main__':
    main()
